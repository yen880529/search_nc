import os
import re
from datetime import datetime
import pandas as pd
from tkinter import *
from tkinter import messagebox
import tkinter as tk
import openpyxl
import time
import webbrowser
import win32com.client as win32
from tkinter import messagebox, simpledialog, font, ttk, Button
import pyperclip
from openpyxl import load_workbook
import json

def get_latest_excel_file(folder_path):
    try:
        # 列出資料夾中的所有檔案
        files = os.listdir(folder_path)

        # 定義符合檔案名稱模式的正則表達式，匹配檔案名類似 "機二組機台排程明細2024.11.04.xlsx"
        pattern = r"機二組機台排程明細(\d{4}\.\d{2}\.\d{2})\.xlsx"

        # 篩選符合正則表達式的檔案
        excel_files = [file for file in files if re.match(pattern, file)]

        # 如果沒有符合條件的檔案，返回 None
        if not excel_files:
            return None

        # 找出符合條件的最新檔案
        latest_file = max(
            excel_files,
            key=lambda f: datetime.strptime(re.search(pattern, f).group(1), "%Y.%m.%d")
        )

        # 返回最新的檔案的完整路徑
        return os.path.join(folder_path, latest_file)

    except Exception as e:
        print(f"Error in get_latest_excel_file: {e}")
        return None


# 更新的 search_excel 函數，使用最新的檔案
def search_excel(search_term1, search_type1, search_term2=None, search_type2=None):

    with open('folders.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 從 JSON 讀取 ball_excel 路徑
    folder_path = data['plan_folders'][0]  # 生產排程

    try:
        # 獲取最新的 Excel 檔案路徑
        latest_file = get_latest_excel_file(folder_path)

        # 確保成功獲取到最新的檔案
        if not latest_file:
            messagebox.showerror("錯誤", "沒有找到符合條件的 Excel 檔案。")
            return []

        print(f"最新檔案: {latest_file}")  # 印出最新檔案名稱，供排查使用

        # 讀取最新的 Excel 檔案，指定讀取第4個工作表（索引為3）
        df = pd.read_excel(latest_file, sheet_name=2)
        
        # 清理列名
        df.columns = df.columns.str.strip().str.lower()

        # 定義搜尋的列名對應
        search_columns = {
            "批號": "批號",
            "料號": "料號",
            "工作中心": "工作中心",
            "工單號碼" : "工單號碼",
            "製程" : "製程",
            "品名" : "品名",
            "圖號": "圖號",
        }
        search_columns = {k: v.strip().lower() for k, v in search_columns.items()}

        # 確認第一行搜尋類型對應的欄位存在
        search_column_name1 = search_columns.get(search_type1.lower())
        if not search_column_name1 or search_column_name1 not in df.columns:
            messagebox.showerror("錯誤", f"找不到指定的欄位：{search_column_name1}")
            return []

        # 第一條件篩選
        if search_term1 == "all":
            search_columns_in_range = ['T51', 'T22', 'L50', 'L52' , 'L40' , 'L43' , 'M41','M42' , 'B11', 'B12', 'R20']
            filtered_df = df[df[search_column_name1].isin(search_columns_in_range)]
        else:
            filtered_df = df[df[search_column_name1].astype(str).str.contains(search_term1, case=False, na=False)]

        # 第二條件篩選（如有）
        if search_term2 and search_type2:
            search_column_name2 = search_columns.get(search_type2.lower())
            if not search_column_name2 or search_column_name2 not in df.columns:
                messagebox.showerror("錯誤", f"找不到指定的欄位：{search_column_name2}")
                return []

            if search_term2 == "all":
                search_columns_in_range = ['T51', 'T22', 'L50', 'L52' , 'L40' , 'L43' , 'M41','M42' , 'B11', 'B12', 'R20']
                filtered_df = filtered_df[filtered_df[search_column_name2].isin(search_columns_in_range)]
            else:
                filtered_df = filtered_df[filtered_df[search_column_name2].astype(str).str.contains(search_term2, case=False, na=False)]

        # 排除不必要的欄位
        exclude_columns = ["作業", "工單數量", "良品數","工時", "預計開始日", "預計完成日", "實際開始日", "實際完成日", "工單狀況", "申請人", "採購單號"]
        filtered_df = filtered_df.drop(columns=[col for col in exclude_columns if col in filtered_df.columns], errors='ignore')

        # 返回篩選後的結果
        return filtered_df.to_dict(orient="records")

    except Exception as e:
        messagebox.showerror("錯誤", f"開啟檔案失敗: {e}")
        return []


# 搜尋按鈕點擊事件
def on_search(search_option1, search_entry1, search_option2, search_entry2, result_tree):
    search_term1 = search_entry1.get().strip()
    search_type1 = search_option1.get()
    search_term2 = search_entry2.get().strip()
    search_type2 = search_option2.get() if search_term2 else None

    if not search_term1:
        messagebox.showwarning("輸入錯誤", "請輸入第一行搜尋詞。")
        return

    # 執行搜尋
    results = search_excel(search_term1, search_type1, search_term2, search_type2)

    # 清空結果窗口
    result_tree.delete(*result_tree.get_children())

    # 顯示搜尋結果
    for row in results:
        # 格式化每一列的數據
        formatted_values = [
            0 if pd.isna(value) else (int(value) if isinstance(value, float) and value.is_integer() else value)
            for value in row.values()
        ]
        result_tree.insert("", "end", values=formatted_values)  # 插入格式化的值

    if not results:
        messagebox.showinfo("無結果", "找不到匹配的資料。")

# 複製選中的行到剪貼簿
def copy_selected_rows(tree):
    selected_items = tree.selection()  # 獲取所有選中的行
    if selected_items:
        # 儲存所有選中行的數據
        rows_data = []
        for item in selected_items:
            row_data = tree.item(item)['values']  # 取得每行的數據
            # 將行數據轉為字符串格式，每列之間用 tab 分隔
            row_text = '\t'.join(map(str, row_data))
            rows_data.append(row_text)
        
        # 將所有行的數據連接成一個字符串，每行之間用換行符分隔
        all_rows_text = '\n'.join(rows_data)
        
        # 將結果複製到剪貼簿
        pyperclip.copy(all_rows_text)
        messagebox.showinfo("複製成功", f"已將 {len(selected_items)} 行複製到剪貼簿！")

# 在右鍵菜單中添加 "複製" 選項
def on_right_click(event, tree):
    selected_item = tree.selection()
    if selected_item:
        # 創建右鍵菜單
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="複製", command=lambda: copy_selected_rows(tree))  # 複製選中的多行
        menu.post(event.x_root, event.y_root)  # 顯示菜單

# 設置 Treeview 的右鍵菜單
def set_right_click_menu(tree):
    tree.bind("<Button-3>", lambda event: on_right_click(event, tree))

def show_comparison_results(comparison_results):
    # 創建新視窗來顯示比對結果
    result_window = tk.Toplevel()
    result_window.title("搜尋結果")
    
    # 創建 Treeview 表格
    tree = ttk.Treeview(result_window, columns=("sheet", "a_col", "b_col", "c_col","e_col","f_col","i_col","k_col","part"), show="headings")
    tree.heading("sheet", text="工作表")
    tree.heading("a_col", text="程式檔號")
    tree.heading("b_col", text="機種")
    tree.heading("c_col", text="名稱")
    tree.heading("e_col", text="圖號")
    tree.heading("f_col", text="版次")
    tree.heading("part", text="料號")
    tree.heading("i_col", text="備註")
    tree.heading("k_col", text="回傳")
    
    

    # 設置欄位的寬度和文字對齊
    tree.column("sheet", anchor="center", width=100)
    tree.column("a_col", anchor="center", width=100)
    tree.column("b_col", anchor="center", width=100)
    tree.column("c_col", anchor="center", width=100)
    tree.column("e_col", anchor="center", width=100)
    tree.column("f_col", anchor="center", width=100)
    tree.column("i_col", anchor="center", width=100)
    tree.column("k_col", anchor="center", width=100)
    tree.column("part", anchor="center", width=150)
    # 插入比對結果資料到表格中
    for sheet, a_col, b_col, c_col, e_col, f_col, i_col, k_col, part, m_col in comparison_results:
            tree.insert("", "end", values=(sheet, a_col, b_col, c_col, e_col, f_col, i_col, k_col, part))
    
    
    # 將 Treeview 滾動條配置
    scrollbar = ttk.Scrollbar(result_window, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    
    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")
    
    # 設置窗口的網格佈局
    result_window.grid_rowconfigure(0, weight=1)
    result_window.grid_columnconfigure(0, weight=1)
    

def extract_numeric_part(part):
    # 若 part 不是字串，先將其轉換為字串
    part = str(part) if not isinstance(part, str) else part
    return ''.join(re.findall(r'\d+', part))

def compare_with_excel(result_tree):
    with open('folders.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 從 JSON 讀取 ball_excel 路徑
    comparison_file_path = data['program_number_2'][0]  # 機二組程式編號

    try:
        print("正在打開 Excel 檔案...")
        workbook = load_workbook(comparison_file_path, data_only=True)
        print("Excel 檔案已打開。")

        # 儲存比對結果
        comparison_results = []

        # 從 result_tree 中取得搜尋結果的料號列表
        print("正在從 result_tree 提取料號...")
        search_result_parts = []
        for item in result_tree.get_children():
            part_number = result_tree.item(item)['values'][2]
            if part_number not in [0, ""]:
                numeric_part = extract_numeric_part(part_number)
                search_result_parts.append(numeric_part)
                print(f"提取料號: {part_number} -> 數字部分: {numeric_part}")

        # 遍歷每個工作表，搜尋符合的資料
        for sheet_name in workbook.sheetnames:
            print(f"正在處理工作表: {sheet_name}")
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 檢查 row 是否包含足夠的列
                if len(row) > 6 and row[6] is not None:  # 確保有第七列且不為空
                    part_number_in_excel = row[6]
                    numeric_part_in_excel = extract_numeric_part(part_number_in_excel)
                    if any(numeric_part_in_excel == part for part in search_result_parts):
                        # 將符合的資料加入比對結果
                        comparison_results.append((
                            sheet_name, row[0] if len(row) > 0 else None, 
                            row[1] if len(row) > 1 else None, 
                            row[2] if len(row) > 2 else None, 
                            row[4] if len(row) > 4 else None, 
                            row[5] if len(row) > 5 else None, 
                            row[8] if len(row) > 8 else None, 
                            row[10] if len(row) > 10 else None, 
                            row[6],
                            row[12] if len(row) > 12 else None  # M 列的值
                        ))
                        print(f"找到匹配: {part_number_in_excel}")

        # 關閉工作簿
        workbook.close()

        # 顯示比對結果
        if comparison_results:
            show_comparison_results(comparison_results)

            # 將比對到的結果行標記為顏色
            for item in result_tree.get_children():
                part_number = result_tree.item(item)['values'][2]  # 料號所在列
                numeric_part = extract_numeric_part(part_number)
                for sheet, a_col, b_col , c_col , e_col , f_col , i_col , k_col, part, m_col in comparison_results:
                    if numeric_part == extract_numeric_part(part):
                        # 根據 M 列是否為空白設置標籤
                        if m_col in [None, ""]:
                            result_tree.item(item, tags=('unmatched',))
                        else:
                            result_tree.item(item, tags=('matched',))
                        break

            # 設置匹配行的樣式
            result_tree.tag_configure('matched', background='lightgreen')
            result_tree.tag_configure('unmatched', background='yellow')

        else:
            messagebox.showinfo("搜尋結果", "未找到符合的資料。")

    except Exception as e:
        print(f"發生錯誤: {e}")
        messagebox.showerror("錯誤", f"比對過程中發生錯誤: {e}")



def open_total_plan_search(root):
    search_window = tk.Toplevel(root)
    search_window.title("Excel 搜尋工具")
    search_window.configure(bg="#DECDF5")  # 設置背景顏色

    with open('folders.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 從 JSON 讀取 ball_excel 路徑
    folder_path = data['plan_folders'][0]  # 生產排程
    # 獲取最新的 Excel 檔案日期
    latest_file = get_latest_excel_file(folder_path)
    latest_date = ""
    if latest_file:
        match = re.search(r"(\d{4}\.\d{2}\.\d{2})", latest_file)
        if match:
            latest_date = match.group(1)
    
    # 使用內建字體
    default_font = font.Font(family="Microsoft JhengHei", size=10) 
    
    # 搜尋條件下拉菜單
    tk.Label(search_window, text="搜尋類型(category)-1:", bg="#DECDF5", font=default_font).grid(row=0, column=0, padx=10, pady=10, sticky="ew")
    search_option1 = tk.StringVar(value="批號")  # 默認選項
    search_dropdown1 = ttk.Combobox(search_window, textvariable=search_option1, values=["批號", "料號", "工單號碼" , "品名" , "工作中心"],
                                    font=default_font, width=15)
    search_dropdown1.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    # 搜尋輸入框
    tk.Label(search_window, text="搜尋名稱(search name)-1:", bg="#DECDF5", font=default_font).grid(row=0, column=2, padx=10, pady=10, sticky="w")
    search_entry1 = tk.Entry(search_window, font=default_font)
    search_entry1.grid(row=0, column=3, padx=10, pady=10, sticky="ew")


     # 第二行搜尋條件下拉菜單和輸入框
    tk.Label(search_window, text="搜尋類型(category)-2:", bg="#DECDF5", font=default_font).grid(row=1, column=0, padx=10, pady=10, sticky="ew")
    search_option2 = tk.StringVar(value="工作中心")  # 默認選項
    search_dropdown2 = ttk.Combobox(search_window, textvariable=search_option2, values=["批號", "料號", "工單號碼" , "品名" , "工作中心"],
                                    font=default_font, width=15)
    search_dropdown2.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

    tk.Label(search_window, text="搜尋名稱(search name)-2:", bg="#DECDF5", font=default_font).grid(row=1, column=2, padx=10, pady=10, sticky="w")
    search_entry2 = tk.Entry(search_window, font=default_font)
    search_entry2.grid(row=1, column=3, padx=10, pady=10, sticky="ew")

    # 搜尋按鈕
    def create_button(text, command, row, column, sticky):
        button = Button(
            search_window,
            text=text,
            command=command,
            bg="#7D6BCB",  # 綠色背景
            fg="white",    # 白色字體
            font=("Microsoft JhengHei", 10, "bold"),  # 調整字體
            borderwidth=2,
            width = 8,
            relief="groove"
        )

        # 鼠標懸停效果
        def on_enter(event):
            button['bg'] = '#ebf2fa'  # 懸停顏色
            button['fg'] = "black"

        def on_leave(event):
            button['bg'] = '#7D6BCB'  # 恢復顏色
            button['fg'] = "white"

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
        button.grid(row=row, column=column, padx=10, pady=5, sticky=sticky)

        return button

    # 搜尋按鈕
    create_button(
        text="搜尋",
        command=lambda: on_search(search_option1, search_entry1, search_option2, search_entry2, result_tree),
        row=0,
        column=4,
        sticky="e"
    )

    # 比對按鈕
    create_button(
        text="程式",
        command=lambda: compare_with_excel(result_tree),
        row=1,
        column=4,
        sticky="e"
    )
    # 結果展示表格
    style = ttk.Style()
    style.configure("Treeview.Heading", anchor="center")  # 標題居中
    style.configure("Treeview", rowheight=25)  # 可選，設定每行高度

    columns = ("批號", "工單號碼", "料號", "品名", "製程", "工作中心" ,"標準工時" , "圖號" , "版次")
    result_tree = ttk.Treeview(search_window, columns=columns, show="headings", height=10)

    for col in columns:
        result_tree.heading(col, text=col)
        result_tree.column(col, anchor="center", width=100)

    result_tree.grid(row=2, column=0, columnspan=5, padx=10, pady=10)

        # 創建 Treeview 並設置列
    result_tree = ttk.Treeview(search_window, columns=columns, show="headings", height=10)

    for col in columns:
        result_tree.heading(col, text=col)
        result_tree.column(col, anchor="center", width=100)

    # 在網格中放置表格
    result_tree.grid(row=2, column=0, columnspan=5, padx=(10, 0), pady=10, sticky="nsew")

    # 創建並配置垂直滾動條
    scrollbar = ttk.Scrollbar(search_window, orient="vertical", command=result_tree.yview)
    result_tree.configure(yscroll=scrollbar.set)

    # 將滾動條放置在表格右側
    scrollbar.grid(row=2, column=5, sticky="nsw", padx=(0, 10), pady=10)

    # 設置右鍵菜單
    set_right_click_menu(result_tree)

    tk.Label(search_window, text=f"最新檔案日期：{latest_date}", bg="#DECDF5", font=default_font, fg="#333").grid(row=3, column=2, columnspan=1, padx=10, pady=10, sticky="ew")
    
    # 未完成標籤
    canvas = tk.Canvas(search_window, width=20, height=20, bg="#DECDF5", highlightthickness=0)
    canvas.create_rectangle(0, 0, 20, 20, fill="yellow", outline="")
    canvas.create_text(30, 10, text="程式未完成", anchor="w", font=("Microsoft JhengHei", 10), fill="#333")
    canvas.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

    # 完成標籤
    canvas = tk.Canvas(search_window, width=20, height=20, bg="#DECDF5", highlightthickness=0)
    canvas.create_rectangle(0, 0, 20, 20, fill="green", outline="")
    canvas.create_text(30, 10, text="程式完成", anchor="w", font=("Microsoft JhengHei", 10), fill="#333")
    canvas.grid(row=3, column=1, padx=10, pady=10, sticky="ew")

    # 綁定單擊事件到圖號列
    result_tree.bind("<Double-1>", lambda event: on_item_click(event, result_tree))
    #search_window.mainloop()
#open_total_plan_search(tk.Tk()) 

def on_item_click(event, tree):
    # 獲取選中的行
    selected_item = tree.selection()
    if selected_item:
        item = tree.item(selected_item)
        # 假設圖號是第五列（索引從0開始）
        圖號 = item['values'][7]
        版本 = item['values'][8]
        
        # 如果版本為空白或 None，將版本設為 "00"
        if not 版本 or 版本 == "None":
            版本 = "00"
        else:
            # 如果版本小於 10，補充 "0" 在前
            版本 = f"{int(版本):02d}"

        # 創建URL
        #url = f"http://192.168.13.74:8080/Work/Graph.aspx?no={圖號}&ver={版本}"
        url = f"sesdap01/Searchpdf/showPDF.aspx?Root={圖號}-{版本}.pdf"
        priority_url = f"http://sesdap01/Searchpdf/Search.aspx"

        # 檢查圖號是否為空白
        if not 圖號 or 圖號 == "None":
            messagebox.showerror("錯誤", "圖號為空白，無法開啟圖面")
        else:
            # 彈出確認提示框
            user_choice = messagebox.askyesno("開啟圖面", f"確定要開啟圖號 {圖號} 的圖面嗎？\n選擇 '否' 手動輸入版本。")

            if user_choice:  # 用戶選擇開啟網址
                webbrowser.open(priority_url)
                time.sleep(1)
                webbrowser.open(url)
            else:  # 用戶選擇手動輸入版本
                # 提示用戶輸入版本
                version_input = simpledialog.askstring("輸入版本", "請輸入版本號（例如 1、2 等）：", parent=tree)

                if version_input is not None:
                    try:
                        # 確保版本是整數，並格式化
                        version_input = str(int(version_input)).zfill(2)
                        # 使用手動輸入的版本創建URL
                        #custom_url = f"http://192.168.13.74:8080/Work/Graph.aspx?no={圖號}&ver={version_input}"
                        custom_url = f"sesdap01/Searchpdf/showPDF.aspx?Root={圖號}-{version_input}.pdf"

                        #一定要開這個
                        priority_url = f"http://sesdap01/Searchpdf/Search.aspx"
                        #呼叫
                        webbrowser.open(priority_url)  # 開啟網址
                        time.sleep(1)
                        webbrowser.open(custom_url)
                    except ValueError:
                        messagebox.showerror("錯誤", "請輸入有效的版本號！")


