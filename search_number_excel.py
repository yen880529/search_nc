import time
import webbrowser
import openpyxl
import win32com.client as win32
import tkinter as tk
from tkinter import messagebox, simpledialog, font, ttk, Button
import json

# 搜尋 Excel 文件中的指定列並返回匹配行
def search_excel(search_term1, search_type1, search_term2=None, search_type2=None):

    with open('folders.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 從 JSON 讀取 ball_excel 路徑
    number_excel = data['program_number_2'][0]  # 機二組程式編號
    
    try:
        # 加載 Excel 文件
        wb = openpyxl.load_workbook(number_excel)
        results = []
        # 依序搜尋工作表
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]  # 第一行的標題
            if search_type1 not in headers:
                continue  # 若找不到指定列名，跳過

            # 定位要搜尋的列
            col_index1 = headers.index(search_type1)
            col_index2 = headers.index(search_type2) if search_type2 in headers else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                cell_value1 = row[col_index1]
                cell_value2 = row[col_index2] if col_index2 is not None else None

                if (cell_value1 and search_term1.lower() in str(cell_value1).lower()) and \
                    (not search_term2 or (cell_value2 and search_term2.lower() in str(cell_value2).lower())):
                    result_row = {headers[i]: row[i] for i in range(len(headers)) if headers[i] not in ["批號", "型號", "作者"]}
                    results.append((sheet, result_row))
        return results
    except Exception as e:
        messagebox.showerror("錯誤", f"開啟檔案失敗: {e}")
        return []

# 搜尋按鈕點擊事件
def on_search(search_option1, search_entry1, search_option2, search_entry2, result_tree):
    search_term1 = search_entry1.get().strip()
    search_type1 = search_option1.get()
    search_term2 = search_entry2.get().strip()
    search_type2 = search_option2.get() if search_term2 else None

    if not search_term1:
        messagebox.showwarning("輸入錯誤", "請輸入搜尋詞。")
        return

    # 執行搜尋
    results = search_excel(search_term1, search_type1, search_term2, search_type2)

    # 清空結果窗口
    result_tree.delete(*result_tree.get_children())

    # 顯示搜尋結果
    for sheet, row in results:
        result_tree.insert("", "end", values=(sheet, *row.values()))  # 展示字典的值

    if not results:
        messagebox.showinfo("無結果", "找不到匹配的資料。")

def open_excel_search(root):
    search_window = tk.Toplevel(root)
    search_window.title("Excel 搜尋工具")
    search_window.configure(bg="#dbe4ee")  # 設置背景顏色
    
    # 使用內建字體
    default_font = font.Font(family="Microsoft JhengHei", size=10) 
    
    # 搜尋條件下拉菜單
    tk.Label(search_window, text="搜尋類型(category)-1:", bg="#dbe4ee", font=default_font).grid(row=0, column=0, padx=10, pady=10, sticky="ew")
    search_option1 = tk.StringVar(value="機種")  # 默認選項
    search_dropdown1 = ttk.Combobox(search_window, textvariable=search_option1, values=["程式", "機種", "名稱", "圖號", "件號"],
                                    font=default_font, width=15)
    search_dropdown1.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    # 搜尋輸入框
    tk.Label(search_window, text="搜尋名稱(search name)-1:", bg="#dbe4ee", font=default_font).grid(row=0, column=2, padx=10, pady=10, sticky="w")
    search_entry1 = tk.Entry(search_window, font=default_font)
    search_entry1.grid(row=0, column=3, padx=10, pady=10, sticky="ew")


     # 第二行搜尋條件下拉菜單和輸入框
    tk.Label(search_window, text="搜尋類型(category)-2:", bg="#dbe4ee", font=default_font).grid(row=1, column=0, padx=10, pady=10, sticky="ew")
    search_option2 = tk.StringVar(value="名稱")  # 默認選項
    search_dropdown2 = ttk.Combobox(search_window, textvariable=search_option2, values=["程式", "機種", "名稱", "圖號", "件號"],
                                    font=default_font, width=15)
    search_dropdown2.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

    tk.Label(search_window, text="搜尋名稱(search name)-2:", bg="#dbe4ee", font=default_font).grid(row=1, column=2, padx=10, pady=10, sticky="w")
    search_entry2 = tk.Entry(search_window, font=default_font)
    search_entry2.grid(row=1, column=3, padx=10, pady=10, sticky="ew")




    # 按鈕樣式函數
# 搜尋按鈕
    def search_button(text, command, row, column, sticky):
        button = Button(
            search_window,
            text=text,
            command=command,
            bg="#427aa1",  # 使用綠色背景
            fg="white",    # 白色字體
            font=("Microsoft JhengHei", 10, "bold"),  # 調整字體
            borderwidth=2,
            width = 8,
            relief="groove"
        )

        # 鼠標懸停效果
        def on_enter(event):
            button['bg'] = '#ebf2fa'  # 懸停顏色
            button['fg'] = "black"

        def on_leave(event):
            button['bg'] = '#427aa1'  # 恢復顏色
            button['fg'] = "white"
            
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
        button.grid(row=row, column=column, padx=10, pady=5, sticky=sticky)

        return button

    # 搜尋按鈕
    search_button(
        text="搜尋",
        command=lambda: on_search(search_option1, search_entry1, search_option2, search_entry2, result_tree),
        row=0,
        column=4,
        sticky="e"
    )



    # 結果展示表格
    style = ttk.Style()
    style.configure("Treeview.Heading", anchor="center")  # 標題居中
    style.configure("Treeview", rowheight=25)  # 可選，設定每行高度

    # 結果展示表格
    columns = ("工作表", "程式", "機種", "名稱", "圖號", "版本" , "件號", "備註", "回傳日")
    result_tree = ttk.Treeview(search_window, columns=columns, show="headings", height=10)

    # 設置每列的標題和寬度，並且內容居中
    for col in columns:
        result_tree.heading(col, text=col)
        result_tree.column(col, anchor="center", width=100)  # 將內容設置為居中

    result_tree.grid(row=2, column=0, columnspan=5, padx=10, pady=10)

    # 創建 Treeview 並設置列
    result_tree = ttk.Treeview(search_window, columns=columns, show="headings", height=10)

    for col in columns:
        result_tree.heading(col, text=col)
        result_tree.column(col, anchor="center", width=100)

    # 在網格中放置表格
    result_tree.grid(row=2, column=0, columnspan=5, padx=(10, 0), pady=10, sticky="nsew")

    # 創建並配置垂直滾動條
    scrollbar = ttk.Scrollbar(search_window, orient="vertical", command=result_tree.yview)
    result_tree.configure(yscroll=scrollbar.set)

    # 將滾動條放置在表格右側
    scrollbar.grid(row=2, column=5, sticky="nsw", padx=(0, 10), pady=10)


    # 綁定單擊事件到圖號列
    result_tree.bind("<Double-1>", lambda event: on_item_click(event, result_tree))

def on_item_click(event, tree):
    # 獲取選中的行
    selected_item = tree.selection()
    if selected_item:
        item = tree.item(selected_item)
        # 假設圖號是第五列（索引從0開始）
        圖號 = item['values'][4]
        版本 = item['values'][5]
        
        # 如果版本為空白或 None，將版本設為 "00"
        if not 版本 or 版本 == "None":
            版本 = "00"
        else:
            # 如果版本小於 10，補充 "0" 在前
            版本 = f"{int(版本):02d}"

        # 創建URL
        #url = f"http://192.168.13.74:8080/Work/Graph.aspx?no={圖號}&ver={版本}"
        url = f"sesdap01/Searchpdf/showPDF.aspx?Root={圖號}-{版本}.pdf"
        priority_url = f"http://sesdap01/Searchpdf/Search.aspx"

        # 檢查圖號是否為空白
        if not 圖號 or 圖號 == "None":
            messagebox.showerror("錯誤", "圖號為空白，無法開啟圖面")
        else:
            # 彈出確認提示框
            user_choice = messagebox.askyesno("開啟圖面", f"確定要開啟圖號 {圖號} 的圖面嗎？\n選擇 '否' 手動輸入版本。")

            if user_choice:  # 用戶選擇開啟網址
                webbrowser.open(priority_url)
                time.sleep(1)
                webbrowser.open(url)
            else:  # 用戶選擇手動輸入版本
                # 提示用戶輸入版本
                version_input = simpledialog.askstring("輸入版本", "請輸入版本號（例如 1、2 等）：", parent=tree)

                if version_input is not None:
                    try:
                        # 確保版本是整數，並格式化
                        version_input = str(int(version_input)).zfill(2)
                        # 使用手動輸入的版本創建URL
                        #custom_url = f"http://192.168.13.74:8080/Work/Graph.aspx?no={圖號}&ver={version_input}"
                        custom_url = f"sesdap01/Searchpdf/showPDF.aspx?Root={圖號}-{version_input}.pdf"

                        #一定要開這個
                        priority_url = f"http://sesdap01/Searchpdf/Search.aspx"
                        #呼叫
                        webbrowser.open(priority_url)  # 開啟網址
                        time.sleep(1)
                        webbrowser.open(custom_url)
                    except ValueError:
                        messagebox.showerror("錯誤", "請輸入有效的版本號！")